from app import db, bcrypt
from app.app_admin.forms import UploadUsersForm
from app.users.models import User, UserInvitation
from flask import Blueprint, redirect, url_for, flash, current_app, json
from flask_login import current_user
from flask_admin import AdminIndexView, BaseView, expose
from flask_admin.contrib.sqla import ModelView
from app.utils import is_valid_phone_number, is_valid_email
import openpyxl
import os
import uuid


app_admin = Blueprint('admin', __name__, template_folder='templates', root_path='admin')


class WrongHeadersError(Exception):

    def __init__(self, message):
        self.message = message


class AuthenticatedAdminIndexView(AdminIndexView):
    def is_accessible(self):
        return current_user.is_authenticated and current_user.has_role('admin')


class AuthenticatedModelView(ModelView):
    def is_accessible(self):
        return current_user.is_authenticated and current_user.has_role('admin')


class AuthenticatedBaseView(BaseView):
    def is_accessible(self):
        return current_user.is_authenticated and current_user.has_role('admin')


class UserView(AuthenticatedModelView):
    column_hide_backrefs = False
    named_filter_urls = True
    form_display_pk = True
    column_display_pk = True
    column_list = ['id', 'first_name', 'last_name', 'email', 'inviter', 'mobile_phone', 'is_active', 'is_bulk_invite']
    form_excluded_columns = ['password']
    # column_searchable_list = ['first_name', 'last_name', 'email']
    column_filters = ['email', 'first_name', 'last_name', 'inviter.first_name', 'inviter.last_name']
    page_size = 50


# class ProcessUserListView(AuthenticatedBaseView):
#     @expose('/')
#     def index(self):
#         if request.args.get('submit_button') == 'Yes':
#             invited_users = UserInvitation.query.filter(UserInvitation.invitation_sent.is_(None)).all()
#
#         else:
#             return redirect(url_for('admin.index'))
#
#     def is_visible(self):
#         return False


def check_if_workbook_ok(excel_file):
    workbook = openpyxl.open(excel_file)
    sheets = workbook.sheetnames
    if len(sheets) > 1:
        raise Exception('The file has more than one sheet. Only one sheet is supported.')
    sheet = workbook[sheets[0]]
    rows_iter = sheet.iter_rows()
    try:
        headers_interim = [cell.value for cell in next(rows_iter)[:4]]
        headers = [header.lower().replace(' ', '_').replace('-', '') for header in headers_interim]
    except AttributeError as error:
        raise AttributeError('Your data does not contain valid headers. '
                             'Please read the instructions carefully')
    for header in headers:
        if header not in ['first_name', 'last_name', 'mobile_phone', 'email']:
            raise Exception('The column headers are not correct, please read the instructions carefully.')
    return workbook


def verify_invitation(data_dict, submitted_number_set, submitted_email_set):

    data_dict['has_errors'] = False
    data_dict['errors'] = []
    data_dict['is_duplication_in_submitted_data'] = False
    data_dict['is_duplicate_of_previous_invitation'] = False
    data_dict['is_for_user_already_in_database'] = False
    data_dict['is_invitation_that_already_in_database'] = False
    data_dict['is_empty_line'] = False

    is_empty_line = True

    # Validate first name
    if not data_dict['first_name']:
        data_dict['errors'].append('The record does not have a first name')

    # Validate last name
    if not data_dict['last_name']:
        data_dict['errors'].append('The record does not have a last name')

    # Validate mobile phone
    if data_dict['mobile_phone']:
        is_empty_line = False
        submitted_number = str(data_dict['mobile_phone'])
        if submitted_number.startswith('264'):
            submitted_number = '+' + submitted_number
            data_dict['mobile_phone'] = submitted_number
        if not is_valid_phone_number(submitted_number):
            data_dict['errors'].append('The mobile phone number does not appear to be valid.')
            data_dict['has_errors'] = True

        user = User.query.filter_by(mobile_phone=data_dict['mobile_phone']).first()
        if user:
            data_dict['errors'].append(f"A user with mobile phone number {data_dict['mobile_phone']} "
                                       f"already exists in the database.")
            data_dict['is_for_user_already_in_database'] = True

        user_invitation = UserInvitation.query.filter_by(mobile_phone=data_dict['mobile_phone']).first()
        if user_invitation:
            data_dict['errors'].append(f"There is already a previous invitation for user with mobile "
                                       f"phone number {data_dict['mobile_phone']} ")
            data_dict['is_duplicate_of_previous_invitation'] = True

        if data_dict['mobile_phone'] in submitted_number_set:
            data_dict['errors'].append(f"Mobile phone {data_dict['mobile_phone']} is already associated with another "
                                       f"record in this dataset.")
            data_dict['is_duplication_in_submitted_data'] = True
        else:
            submitted_number_set.add(data_dict['mobile_phone'] )

    # Validate e-mail
    if data_dict['email']:
        is_empty_line = False
        if not is_valid_email(data_dict['email']):
            data_dict['errors'].append('The e-mail address does not appear to be valid.')
            data_dict['has_errors'] = True

        user = User.query.filter_by(email=data_dict['email']).first()
        if user:
            data_dict['errors'].append(f"A user with email {data_dict['email']} "
                                       f"already exists in the database.")
            data_dict['is_for_user_already_in_database'] = True

        user_invitation = UserInvitation.query.filter_by(mobile_phone=data_dict['mobile_phone']).first()
        if user_invitation:
            data_dict['errors'].append(f"There is already a previous invitation for user with email "
                                       f"{data_dict['email']} ")
            data_dict['is_duplicate_of_previous_invitation'] = True

        if data_dict['mobile_phone'] in submitted_email_set:
            data_dict['errors'].append(f"E-mail {data_dict['email']} is already associated with another "
                                       f"record in this dataset.")
            data_dict['is_duplication_in_submitted_data'] = True
        else:
            submitted_email_set.add(data_dict['mobile_phone'] )

    data_dict['is_empty_line'] = is_empty_line
    if len(data_dict['errors']) > 0:
        data_dict['has_errors'] = True


def process_workbook(workbook):

    valid_invitations = {}
    invalid_invitations = {}

    invitation_metadata = {"errors": 0,
                           "has_duplication_in_submitted_data": 0,
                           "is_duplication_of_previous_invitations": 0,
                           "has_users_that_already_exists_in_database": 0,
                           "has_invitations_that_already_exists_in_database": 0}

    sheet = workbook[workbook.sheetnames[0]]
    rows_iter = sheet.iter_rows()

    headers_interim = [cell.value for cell in next(rows_iter)[:4]]
    headers = [header.lower().replace(' ', '_').replace('-', '') for header in headers_interim]
    submitted_numbers_set = set()
    submitted_emails_set = set()

    while True:
        try:
            data = [cell.value for cell in next(rows_iter)[:4]]
            data_dict = dict(zip(headers, data))
            record_id = str(uuid.uuid4())
            verify_invitation(data_dict, submitted_numbers_set, submitted_emails_set)
            if not data_dict['is_empty_line']:
                del data_dict['is_empty_line']
                if data_dict['has_errors']:
                    invitation_metadata['errors'] += 1
                    invalid_invitations[record_id] = data_dict
                else:
                    valid_invitations[record_id] = data_dict
        except StopIteration:
            break

    final_data = {'metadata': invitation_metadata,
                  'valid_invitations': valid_invitations,
                  'invalid_invitations': invalid_invitations}

    return final_data


class UserUploadView(AuthenticatedBaseView):

    @expose('/', methods=('GET', 'POST'))
    def index(self):
        form = UploadUsersForm()
        if form.validate_on_submit():
            show_send = False
            hashed_password = bcrypt.generate_password_hash(form.password.data).decode('utf-8')
            excel_file = form.excel_file.data
            process_errors = False
            try:
                workbook = check_if_workbook_ok(excel_file)
                final_data = process_workbook(workbook)
                if final_data['metadata']['errors'] > 0:
                    file_name = str(uuid.uuid4())
                    temp_path = os.path.join(current_app.root_path, 'static', 'temp', file_name)
                    the_json = json.dumps(final_data)
                    with open(temp_path, 'w') as json_file:
                        json_file.write(the_json)
                    if final_data['metadata']['errors'] > 10:
                        flash(f"There are more than 10 ({final_data['metadata']['errors']}) problems with your file. "
                              f"It needs some cleaning up. Please do that an submit again.", "error")
                    else:
                        flash('There are some issues with your Excel file. Please see below.', 'warning')
                        return redirect(url_for('upload.process_errors', file_name=file_name))
            except Exception as error:
                flash(f"There was a problem during the processing of your file. "
                      f"The error message reads as follows: <em>{error}</em>",
                      'error')
        else:
            flash("You need to upload a file with an '.xlsx' extension")
        return self.render('admin/user_upload.html', form=form)

    @expose('/process_errors/<file_name>', methods=('GET', 'POST'))
    def process_errors(self, file_name):
        temp_path = os.path.join(current_app.root_path, 'static', 'temp', file_name)
        with open(temp_path, 'r') as json_file:
            json_data = json.load(json_file)
        # os.remove(temp_path)
        return self.render('admin/process_excel.html', submitted_data=json.dumps(json_data))

    def is_accessible(self):
        return current_user.is_authenticated


def add_admin_views(admin):
    from app.users.models import User, Role
    admin.add_view(UserView(User, db.session, category='Users'))
    admin.add_view(AuthenticatedModelView(Role, db.session, category='Users'))

    admin.add_view(UserUploadView('Upload users', endpoint='upload', category='Users'))
    admin.add_sub_category(name='User Management', parent_name='Users')

    from app.geography.models import Country, SubdivisionType, Subdivision, PopulatedAreaType, PopulatedArea
    admin.add_view(AuthenticatedModelView(Country, db.session, category='Geo'))
    admin.add_view(AuthenticatedModelView(SubdivisionType, db.session, category='Geo'))
    admin.add_view(AuthenticatedModelView(Subdivision, db.session, category='Geo'))
    admin.add_view(AuthenticatedModelView(PopulatedAreaType, db.session, category='Geo'))
    admin.add_view(AuthenticatedModelView(PopulatedArea, db.session, category='Geo'))

    admin.add_sub_category(name='Geo', parent_name='Geo')
