from openpyxl import load_workbook

wb = load_workbook('/home/lbecker/Documents/Updated Adres Lys - Test.xlsx')

sheet = wb.active

rows = sheet.iter_rows()


while True:
    data = [cell.value for cell in next(rows)]
    print(data)

