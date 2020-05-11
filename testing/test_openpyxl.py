from openpyxl import load_workbook

wb = load_workbook("C:\\Users\\lbecker\\Documents\\aa_maranatha_users.xlsx")

name = wb.sheetnames[0]

sheet = wb[name]

row_iter = sheet.iter_rows()

while True:
    try:
        data = [cell.value for cell in next(row_iter)]
        print(data)
    except StopIteration:
        break